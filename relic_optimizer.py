"""
Relic Calculator v2 — Optimization Engine
==========================================
Chain model: develop seed → swap → develop relay → ... → swap → develop target.

Fragment rules:
  - Chain steps (seed/relay): specific shards first, then only enough universals
    to complete the current leg boundary. No specific shards left on swap.
  - Target step: specific shards first, then maximize universals up to target goal.
  - Post-chain: remaining universals maximized across all undeveloped targets.

Mandatory relics (inter1/inter2): placed in any chain position (seed or relay).
Tool warns if a different choice would give a better result.

Usage:
    py relic_optimizer.py
    py relic_optimizer.py "path/to/Relic_Calculator_v2.xlsx"

Requires:
    pip install openpyxl
"""

import sys
from itertools import combinations, permutations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Star level data ────────────────────────────────────────────────────────────
CUMUL = [
    0, 2, 4, 6, 8, 10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 58, 66,
    74, 82, 90, 98, 106, 114, 122, 130, 142, 154, 166, 178, 190, 202,
    214, 226, 238, 250, 266, 282, 298, 314, 330, 346, 362, 378, 394, 410,
    426, 442, 458, 474, 490, 510, 530, 550, 570, 590, 610, 630, 650, 670,
    690, 716, 742, 768, 794, 820, 848, 876, 904, 932, 960, 988, 1016,
    1044, 1072, 1100, 1128, 1156, 1184, 1212, 1240, 1268, 1296, 1324,
    1352, 1380, 1408, 1436, 1464, 1492, 1520, 1548, 1576, 1604, 1632,
    1660, 1688, 1716, 1744, 1772, 1800,
]

STAR_OPTIONS = [
    "0★",
    "Y★1","Y★2","Y★3","Y★4","Y★5",
    "R★1","R★2","R★3","R★4","R★5",
    "P★1","P★2","P★3","P★4","P★5",
    "B★1","B★2","B★3","B★4","B★5",
]

FULL_LABELS = ["0★ (0/5)"]
for _t in ["Y","R","P","B"]:
    for _s in range(1, 6):
        for _l in range(1, 6):
            FULL_LABELS.append(f"{_t}★{_s} ({_l}/5)")


def star_leg_to_idx(star: str, leg) -> int:
    star = str(star).strip()
    if star in ("0★", "", "—"):
        return 0
    try:
        star_pos = STAR_OPTIONS.index(star)
    except ValueError:
        return 0
    leg_num = int(str(leg).split("/")[0]) if leg and str(leg) not in ("", "—") else 1
    return (star_pos - 1) * 5 + leg_num


def idx_to_star_leg(idx: int) -> tuple:
    if idx <= 0:
        return "0★", "0/5"
    star_num = (idx - 1) // 5
    leg_num  = (idx - 1) % 5 + 1
    return STAR_OPTIONS[star_num + 1], f"{leg_num}/5"


def shards_needed(from_idx: int, to_idx: int) -> int:
    if to_idx <= from_idx:
        return 0
    return CUMUL[to_idx] - CUMUL[from_idx]


def max_level_reachable(current: int, shards: int) -> int:
    best = current
    for lv in range(current + 1, 101):
        if shards_needed(current, lv) <= shards:
            best = lv
        else:
            break
    return best


# ── Relic definitions ──────────────────────────────────────────────────────────
UNIVERSAL_RELICS = [
    "Duke's Signet Ring", "Eternal Wings", "Frost Diadem", "Royalty",
    "War Flag", "Scale of Injustice", "Mighty Gold",
    "Persecution", "Anti-Magic Handcuffs", "Moonstone",
]
PREFERRED_INTER = ["Duke's Signet Ring", "Eternal Wings"]
SETS = {
    "League": ["Petrification Staff", "Soul Guard Orb", "Feather of the Pact"],
    "Horde":  ["Thunder Judgment", "Dragonheart", "Dragonbone Amulet"],
    "Nature": ["Vineborne Bow", "Undefeated Crown", "Sacred Scroll"],
}
ALL_SET_RELICS = [r for relics in SETS.values() for r in relics]
ALL_RELICS     = UNIVERSAL_RELICS + ALL_SET_RELICS

RELIC_NAME_PT = {
    "Duke's Signet Ring":   "Anel Sinete do Duque",
    "Eternal Wings":        "Asas Eternas",
    "Frost Diadem":         "Diadema de Gelo",
    "Royalty":              "Realeza",
    "War Flag":             "Bandeira de Guerra",
    "Scale of Injustice":   "Balança da Injustiça",
    "Mighty Gold":          "Suborno",
    "Persecution":          "Perseguição",
    "Anti-Magic Handcuffs": "Algemas Anti-Magia",
    "Moonstone":            "Pedra da Lua",
    "Petrification Staff":  "Cajado de Petrificação",
    "Soul Guard Orb":       "Orbe Guardião da Alma",
    "Feather of the Pact":  "Pena do Pacto",
    "Thunder Judgment":     "Julgamento Trovejante",
    "Dragonheart":          "Coração de Dragão",
    "Dragonbone Amulet":    "Amuleto de Osso de Dragão",
    "Vineborne Bow":        "Arco de Vinhas",
    "Undefeated Crown":     "Coroa do Invicto",
    "Sacred Scroll":        "Pergaminho Sagrado",
}

# Portrait images stored in relic_imgs/ (extracted from TopHeroes Tables spreadsheet)
RELIC_IMAGES = {
    # Legendary — Universal
    "Duke's Signet Ring":   "dukes_signet_ring.png",
    "Eternal Wings":        "eternal_wings.png",
    "Frost Diadem":         "frost_diadem.png",
    "Royalty":              "royalty.png",
    "War Flag":             "war_flag.png",
    "Scale of Injustice":   "scale_of_injustice.png",
    "Mighty Gold":          "mighty_gold.png",
    "Persecution":          "persecution.png",
    "Anti-Magic Handcuffs": "anti_magic_handcuffs.png",
    "Moonstone":            "moonstone.png",
    # Legendary — League Set
    "Petrification Staff":  "petrification_staff.png",
    "Soul Guard Orb":       "soul_guard_orb.png",
    "Feather of the Pact":  "feather_of_the_pact.png",
    # Legendary — Horde Set
    "Thunder Judgment":     "thunder_judgment.png",
    "Dragonheart":          "dragonheart.png",
    "Dragonbone Amulet":    "dragonbone_amulet.png",
    # Legendary — Nature Set
    "Vineborne Bow":        "vineborne_bow.png",
    "Undefeated Crown":     "undefeated_crown.png",
    "Sacred Scroll":        "sacred_scroll.png",
    # Epic
    "Venom Fang":           "venom_fang.png",
    "Terra Orb":            "terra_orb.png",
    "Vowkeeper":            "vowkeeper.png",
    "Marshal's Warhorn":    "marshals_warhorn.png",
    "Philosopher's Rune":   "philosophers_rune.png",
    "Electric Frog":        "electric_frog.png",
    "Eye of Fear":          "eye_of_fear.png",
    # Rare
    "Elixir Bottle":        "elixir_bottle.png",
    "Balance Brooch":       "balance_brooch.png",
    "Hero's Amulet":        "heros_amulet.png",
}

RARE_RELICS = ["Elixir Bottle", "Balance Brooch", "Hero's Amulet"]

EPIC_RELICS = [
    "Venom Fang", "Terra Orb", "Vowkeeper",
    "Marshal's Warhorn", "Philosopher's Rune", "Electric Frog", "Eye of Fear",
]


# ── Read inventory ─────────────────────────────────────────────────────────────
def read_inventory(filepath: str) -> dict:
    wb      = openpyxl.load_workbook(filepath, data_only=True)
    ws_inv  = wb["📦 Inventory"]
    ws_calc = wb["⚜️ Relic Calculator"]

    universal_shards = int(ws_inv["B3"].value or 0)

    relics = {}
    for row in ws_inv.iter_rows(min_row=7, values_only=True):
        raw_name = str(row[1] or "").strip().lstrip("⭐").strip()
        if raw_name not in ALL_RELICS:
            continue
        star    = str(row[3] or "0★").strip()
        leg_raw = row[4]
        leg_str = str(leg_raw or "—").strip()
        idx     = star_leg_to_idx(star, leg_str)
        spec    = int(row[5] or 0)
        can_use = str(row[6] or "No").strip() == "Yes"
        relics[raw_name] = {
            "star_idx":        idx,
            "specific_shards": spec,
            "can_use":         can_use,
        }

    target_set   = str(ws_calc["C2"].value or "Horde").strip()
    target_relic = str(ws_calc["C3"].value or "").strip()
    priority     = [
        str(ws_calc["E5"].value or "").strip(),
        str(ws_calc["F5"].value or "").strip(),
        str(ws_calc["G5"].value or "").strip(),
    ]
    inter1        = str(ws_calc["C6"].value or PREFERRED_INTER[0]).strip()
    inter2        = str(ws_calc["F6"].value or "").strip()
    tgt_star      = str(ws_calc["C8"].value or "P★5").strip()
    tgt_leg       = str(ws_calc["E8"].value or "5/5").strip()
    target_level  = star_leg_to_idx(tgt_star, tgt_leg)
    hammers_avail = int(ws_calc["C9"].value or 0)

    return {
        "universal_shards": universal_shards,
        "relics":           relics,
        "config": {
            "target_set":    target_set,
            "target_relic":  target_relic,
            "priority":      priority,
            "inter1":        inter1,
            "inter2":        inter2,
            "target_level":  target_level,
            "hammers_avail": hammers_avail,
        },
    }


# ── Fragment development ───────────────────────────────────────────────────────

def chain_develop_options(from_level: int, specific_shards: int, universal_avail: int) -> list:
    """
    Return candidate (level, sp_used, u_used) outcomes for chain development.

    Up to two candidates:
      A) Highest leg boundary reachable with 0 universals (may waste some sp).
      B) First leg boundary that uses ALL specific shards (costs minimum universals).
    Returns one candidate when A == B, or falls back to best-effort when
    neither A nor B is feasible.
    """
    if from_level >= 100 or (specific_shards <= 0 and universal_avail <= 0):
        return [(from_level, 0, 0)]

    from_cumul = CUMUL[from_level]
    candidates = []

    if specific_shards > 0:
        # Option A: highest boundary reachable with 0 universals
        best_0u_lv, best_0u_sp = from_level, 0
        for lv in range(from_level + 1, 101):
            cost = CUMUL[lv] - from_cumul
            if cost <= specific_shards:
                best_0u_lv, best_0u_sp = lv, cost
            else:
                break
        if best_0u_lv > from_level:
            candidates.append((best_0u_lv, best_0u_sp, 0))

        # Option B: first boundary where cost >= sp (uses all sp + min universals)
        for lv in range(from_level + 1, 101):
            cost = CUMUL[lv] - from_cumul
            if cost >= specific_shards:
                u_needed = cost - specific_shards
                if u_needed <= universal_avail and lv != best_0u_lv:
                    candidates.append((lv, specific_shards, u_needed))
                break

    # Fallback: best-effort spending all available resources
    if not candidates:
        total = specific_shards + universal_avail
        best  = from_level
        for lv in range(from_level + 1, 101):
            if CUMUL[lv] - from_cumul <= total:
                best = lv
            else:
                break
        if best > from_level:
            cost    = CUMUL[best] - from_cumul
            sp_used = min(specific_shards, cost)
            u_used  = cost - sp_used
            candidates.append((best, sp_used, u_used))

    return candidates if candidates else [(from_level, 0, 0)]


def chain_develop(from_level: int, specific_shards: int, universal_avail: int) -> tuple:
    """Single-result wrapper — returns the min-universal option from chain_develop_options."""
    opts = chain_develop_options(from_level, specific_shards, universal_avail)
    # Prefer the option with the highest level; break ties by lowest u
    return max(opts, key=lambda o: (o[0], -o[2]))


def develop_max(from_level: int, specific_shards: int, universal_budget: int,
                cap: int = 100) -> tuple:
    """
    Develop a relic using ALL available resources up to `cap`.
    No leg constraint — maximize the final level.

    Used for: target development after receiving chain level, and post-chain
    top-up of all targets.

    Returns (achieved_level, sp_used, u_used).
    """
    if from_level >= cap or from_level >= 100:
        return from_level, 0, 0
    total = specific_shards + universal_budget
    if total <= 0:
        return from_level, 0, 0
    effective_cap = min(cap, 100)
    from_cumul    = CUMUL[from_level]
    best          = from_level
    for lv in range(from_level + 1, effective_cap + 1):
        if CUMUL[lv] - from_cumul <= total:
            best = lv
        else:
            break
    if best == from_level:
        return from_level, 0, 0
    cost    = CUMUL[best] - from_cumul
    sp_used = min(specific_shards, cost)
    u_used  = cost - sp_used
    return best, sp_used, u_used


# ── Chain simulation ───────────────────────────────────────────────────────────

def simulate_chain(
    chain_relics:      list,
    target_name:       str,
    target_orig_level: int,
    target_specific:   int,
    universal_pool:    int,
    target_goal:       int = 100,
) -> tuple:
    """
    Simulate a Miracle Hammer chain, trying all combinations of development
    options (0-universal stop vs min-universal stop) at every relay position.

    chain_relics: [(name, orig_level, specific_shards), ...]
                  Position 0 is the seed; remaining entries are relays.

    Returns (final_target_level, total_u_used, steps), or None if no valid
    chain exists (every path would downgrade the target).
    """
    if not chain_relics:
        return None

    best_lv = -1
    best_u  = float("inf")
    best_st = None
    work    = []   # mutable step buffer; copied only when a new best is found

    def dfs(relic_idx: int, carrier_level: int, carrier_name: str, remaining_u: int):
        nonlocal best_lv, best_u, best_st

        if relic_idx == len(chain_relics):
            # All relays done — final swap to target
            if carrier_level <= target_orig_level:
                return
            swap = {
                "type":     "swap",
                "relic_a":  carrier_name, "a_from": carrier_level,     "a_to": target_orig_level,
                "relic_b":  target_name,  "b_from": target_orig_level, "b_to": carrier_level,
            }
            achieved, sp_used, u_used = develop_max(
                carrier_level, target_specific, remaining_u, cap=target_goal
            )
            dev = {
                "type":    "develop",
                "relic":   target_name,
                "from":    carrier_level, "to":      achieved,
                "sp_used": sp_used,       "u_used":  u_used,
            }
            total_u = (universal_pool - remaining_u) + u_used
            if achieved > best_lv or (achieved == best_lv and total_u < best_u):
                best_lv = achieved
                best_u  = total_u
                best_st = work[:] + [swap, dev]
            return

        name, orig_level, specific = chain_relics[relic_idx]
        incoming = carrier_level
        swap = {
            "type":     "swap",
            "relic_a":  carrier_name, "a_from": carrier_level, "a_to": orig_level,
            "relic_b":  name,         "b_from": orig_level,    "b_to": carrier_level,
        }

        for achieved, sp_used, u_used in chain_develop_options(incoming, specific, remaining_u):
            dev = {
                "type":    "develop",
                "relic":   name,
                "from":    incoming, "to":      achieved,
                "sp_used": sp_used,  "u_used":  u_used,
            }
            work.append(swap)
            work.append(dev)
            dfs(relic_idx + 1, achieved, name, remaining_u - u_used)
            work.pop()
            work.pop()

    # ── Seed development ──────────────────────────────────────────────────────
    seed_name, seed_level, seed_specific = chain_relics[0]

    if seed_specific > 0 and seed_level < 100:
        any_launched = False
        for achieved, sp_used, u_used in chain_develop_options(seed_level, seed_specific, universal_pool):
            if achieved > seed_level:
                dev = {
                    "type":    "develop",
                    "relic":   seed_name,
                    "from":    seed_level, "to":      achieved,
                    "sp_used": sp_used,    "u_used":  u_used,
                }
                work.append(dev)
                dfs(1, achieved, seed_name, universal_pool - u_used)
                work.pop()
                any_launched = True
        if not any_launched:
            dfs(1, seed_level, seed_name, universal_pool)
    else:
        dfs(1, seed_level, seed_name, universal_pool)

    return (best_lv, best_u, best_st) if best_lv >= 0 else None


# ── Per-target chain optimizer ─────────────────────────────────────────────────

def best_chain_for_target(
    available_relics: list,
    mandatory_names:  set,
    n_hammers:        int,
    target_name:      str,
    target_level:     int,
    target_specific:  int,
    universal_pool:   int,
    target_goal:      int,
    first_relay:      str = None,
) -> tuple:
    """
    Find the best chain for this target by trying all valid permutations.

    available_relics: [(name, level, specific), ...] — eligible universals
    mandatory_names:  names that MUST appear somewhere in the chain
    n_hammers:        swaps to use (= number of relics before target)
    first_relay:      if set, this relic is fixed at relay position 0
                      (chain_relics[1]); the seed swaps into it first.

    Returns (best_level, best_u, best_steps), or (-1, inf, None).
    """
    if n_hammers <= 0:
        return -1, float("inf"), None

    best_lv = -1
    best_u  = float("inf")
    best_st = None

    def try_chain(chain):
        nonlocal best_lv, best_u, best_st
        res = simulate_chain(chain, target_name, target_level, target_specific,
                             universal_pool, target_goal)
        if res is None:
            return
        lv, u, steps = res
        if lv > best_lv or (lv == best_lv and u < best_u):
            best_lv, best_u, best_st = lv, u, steps

    def done(): return best_lv >= target_goal

    if first_relay:
        if n_hammers < 2:
            return -1, float("inf"), None
        fr_relic = next((r for r in available_relics if r[0] == first_relay), None)
        if fr_relic is None:
            return -1, float("inf"), None
        rest      = [r for r in available_relics if r[0] != first_relay]
        rest_mand = [r for r in rest if r[0] in mandatory_names]
        rest_opt  = [r for r in rest if r[0] not in mandatory_names]
        rest_opt  = sorted(rest_opt, key=lambda r: r[1] + r[2], reverse=True)[:4]
        n_rest    = n_hammers - 1   # slots for seed + remaining relays
        n_mand_r  = len(rest_mand)
        if n_mand_r > n_rest:
            for sub in combinations(rest_mand, n_rest):
                for pre in permutations(sub):
                    try_chain([pre[0], fr_relic] + list(pre[1:]))
                    if done(): return best_lv, best_u, best_st
        else:
            n_opt_need = n_rest - n_mand_r
            for n_pick in range(min(n_opt_need, len(rest_opt)), -1, -1):
                for opt_c in combinations(rest_opt, n_pick):
                    pool = rest_mand + list(opt_c)
                    if not pool:   # need at least the seed
                        continue
                    for pre in permutations(pool):
                        try_chain([pre[0], fr_relic] + list(pre[1:]))
                        if done(): return best_lv, best_u, best_st
                if best_lv >= 0:
                    break
    else:
        mandatory = [r for r in available_relics if r[0] in mandatory_names]
        optional  = [r for r in available_relics if r[0] not in mandatory_names]
        optional  = sorted(optional, key=lambda r: r[1] + r[2], reverse=True)[:4]
        n_mand    = len(mandatory)
        if n_mand > n_hammers:
            for subset in combinations(mandatory, n_hammers):
                for perm in permutations(subset):
                    try_chain(list(perm))
                    if done(): return best_lv, best_u, best_st
        else:
            n_opt_need = n_hammers - n_mand
            for n_pick in range(min(n_opt_need, len(optional)), -1, -1):
                for opt_combo in combinations(optional, n_pick):
                    pool = mandatory + list(opt_combo)
                    for perm in permutations(pool):
                        try_chain(list(perm))
                        if done(): return best_lv, best_u, best_st
                if best_lv >= 0:
                    break

    return best_lv, best_u, best_st


# ── Route scoring ──────────────────────────────────────────────────────────────

def score_route(
    final_levels:   dict,
    final_specific: dict,
    targets:        list,
    target_goal:    int,
    universal_init: int,
    universal_used: int,
) -> int:
    """
    Priority-weighted sum of capped target levels, including simulated
    post-chain development with remaining universals.
    """
    levels   = dict(final_levels)
    specific = dict(final_specific)
    remaining = universal_init - universal_used
    for t in targets:
        cur = levels.get(t, 0)
        if cur >= target_goal or remaining <= 0:
            continue
        sp = specific.get(t, 0)
        new_lv, sp_use, u_use = develop_max(cur, sp, remaining, cap=target_goal)
        levels[t]   = new_lv
        specific[t] = max(0, sp - sp_use)
        remaining  -= u_use

    score = 0
    n     = len(targets)
    for i, t in enumerate(targets):
        lv = min(levels.get(t, 0), target_goal)
        score += (n - i) * lv
    return score


# ── Enumeration helpers ────────────────────────────────────────────────────────

def _gen_all_splits(total: int, n: int):
    """All non-negative integer tuples of length n summing to total."""
    if n == 1:
        yield (total,)
        return
    for h in range(0, total + 1):
        for rest in _gen_all_splits(total - h, n - 1):
            yield (h,) + rest


def _gen_mandatory_assignments(mandatory_names: list, targets: list, split: tuple):
    """
    Yield all valid ways to assign mandatory relics to target chains.
    Each mandatory is assigned to exactly one target that has ≥ 1 hammer.
    A target receives at most as many mandatories as its hammer count.
    Yields dict {target_name: [mandatory_names...]}.
    """
    if not mandatory_names:
        yield {}
        return

    eligible = [(t, h) for t, h in zip(targets, split) if h > 0]
    if not eligible:
        return

    def assign(idx, current):
        if idx == len(mandatory_names):
            yield {k: list(v) for k, v in current.items() if v}
            return
        mname = mandatory_names[idx]
        for t, h in eligible:
            if len(current.get(t, [])) >= h:
                continue  # target already has as many mandatories as its hammers
            current.setdefault(t, []).append(mname)
            yield from assign(idx + 1, current)
            current[t].pop()
            if not current[t]:
                del current[t]

    yield from assign(0, {})


# ── Configuration evaluator ────────────────────────────────────────────────────

def _evaluate_config(
    targets:        list,
    split:          tuple,
    mand_assign:    dict,
    levels_init:    dict,
    specific_init:  dict,
    universal_init: int,
    target_goal:    int,
    first_relays:   dict = None,
) -> tuple:
    """
    Simulate all chains for one (split, mandatory_assignment) combination.

    first_relays: optional {target_name: relay_name} — fixes the first relay
    position for each target's chain.

    Returns (score, result_dict), or (None, None) if any required chain fails.
    """
    levels    = dict(levels_init)
    specific  = dict(specific_init)
    universal = universal_init
    all_steps = []
    used      = set(targets)
    total_h   = 0
    total_u   = 0

    for target, h in zip(targets, split):
        if h == 0:
            continue

        chain_mand = set(mand_assign.get(target, []))
        available  = [
            (r, levels[r], specific.get(r, 0))
            for r in UNIVERSAL_RELICS
            if r not in used
        ]
        fr = (first_relays or {}).get(target)

        lv, u, steps = best_chain_for_target(
            available, chain_mand, h,
            target, levels[target], specific.get(target, 0),
            universal, target_goal,
            first_relay=fr,
        )

        if steps is None:
            return None, None

        all_steps += steps
        total_h   += h
        total_u   += u
        universal -= u

        for step in steps:
            if step["type"] == "swap":
                used.add(step["relic_a"])            # outgoing carrier = chain relic
                levels[step["relic_a"]] = step["a_to"]
                levels[step["relic_b"]] = step["b_to"]
            elif step["type"] == "develop":
                r = step["relic"]
                levels[r]   = step["to"]
                specific[r] = max(0, specific[r] - step["sp_used"])

    sc = score_route(levels, specific, targets, target_goal, universal_init, total_u)
    return sc, {
        "steps":          all_steps,
        "final_levels":   levels,
        "final_specific": specific,
        "hammers_used":   total_h,
        "universal_used": total_u,
        "targets":        list(targets),
        "assignment":     dict(mand_assign),
        "suboptimal_note": "",
    }


# ── Main route computation ─────────────────────────────────────────────────────

def compute_route(inv: dict) -> dict:
    cfg           = inv["config"]
    target_goal   = cfg["target_level"]
    hammers_avail = cfg["hammers_avail"]

    levels_init    = {r: inv["relics"].get(r, {}).get("star_idx",        0) for r in ALL_RELICS}
    specific_init  = {r: inv["relics"].get(r, {}).get("specific_shards", 0) for r in ALL_RELICS}
    universal_init = inv["universal_shards"]

    # Targets in priority order, skipping those already at goal
    if cfg["target_set"] in SETS:
        set_relics = SETS[cfg["target_set"]]
        priority   = [p for p in cfg["priority"] if p in set_relics]
        ordered    = priority + [r for r in set_relics if r not in priority]
        targets    = [r for r in ordered if levels_init.get(r, 0) < target_goal]
    elif cfg["target_relic"] and cfg["target_relic"] in ALL_RELICS:
        targets = ([cfg["target_relic"]]
                   if levels_init.get(cfg["target_relic"], 0) < target_goal else [])
    else:
        targets = []

    all_targets = targets[:]

    empty = {
        "steps": [], "final_levels": dict(levels_init),
        "final_specific": dict(specific_init),
        "hammers_used": 0, "universal_used": 0,
        "targets": all_targets, "suboptimal_note": "",
    }
    if not targets:
        return empty

    inter1 = cfg.get("inter1", "")
    inter2 = cfg.get("inter2", "")
    inter1 = inter1 if inter1 in UNIVERSAL_RELICS else PREFERRED_INTER[0]
    inter2 = inter2 if inter2 in UNIVERSAL_RELICS else ""
    mandatory_names = [i for i in [inter1, inter2] if i]

    first_relays = {k: v for k, v in cfg.get('first_relays', {}).items() if v}

    def _search(mandatory, fr=None):
        best_sc  = -1
        best_res = None
        for split in _gen_all_splits(hammers_avail, len(targets)):
            for mand_assign in _gen_mandatory_assignments(mandatory, targets, split):
                sc, res = _evaluate_config(
                    targets, split, mand_assign,
                    levels_init, specific_init, universal_init, target_goal,
                    first_relays=fr,
                )
                if res is None:
                    continue
                if sc > best_sc or (sc == best_sc
                                    and res["universal_used"] < best_res["universal_used"]):
                    best_sc, best_res = sc, res
        return best_sc, best_res

    def _post_chain(result):
        """Apply post-chain universal development in-place and return the result."""
        fl        = result["final_levels"]
        fs        = result["final_specific"]
        remaining = universal_init - result["universal_used"]
        for relic in all_targets:
            if remaining <= 0:
                break
            cur = fl.get(relic, 0)
            if cur >= target_goal:
                continue
            sp = fs.get(relic, 0)
            new_lv, sp_use, u_use = develop_max(cur, sp, remaining, cap=target_goal)
            if new_lv > cur:
                result["steps"].append({
                    "type": "develop", "relic": relic,
                    "from": cur, "to": new_lv,
                    "sp_used": sp_use, "u_used": u_use,
                })
                fl[relic]  = new_lv
                fs[relic]  = max(0, sp - sp_use)
                remaining -= u_use
                result["universal_used"] += u_use
        return result

    best_score, best_result = _search(mandatory_names, first_relays or None)

    if best_result is None:
        return empty

    _post_chain(best_result)

    # ── Suboptimality check ───────────────────────────────────────────────────
    # When first_relays is set: compare against mandatory-only (no first_relay
    # restriction) — shows whether the position constraint is costing efficiency.
    # When only mandatory: compare against fully free — shows whether the
    # mandatory choice is suboptimal.
    optimal_result  = None
    suboptimal_note = ""

    if first_relays:
        cmp_score, cmp_result = _search(mandatory_names)
        cmp_label = "primeira relay fixada por alvo"
    else:
        cmp_score, cmp_result = _search([])
        cmp_label = (f"relays obrigatórios ({', '.join(mandatory_names)})"
                     if mandatory_names else "configuração atual")

    if cmp_result is not None:
        _post_chain(cmp_result)
        _is_sub = (
            cmp_score > best_score or
            (cmp_score == best_score and
             cmp_result["universal_used"] < best_result["universal_used"])
        )
        if _is_sub:
            suboptimal_note = (
                f"Atenção: {cmp_label} não é a escolha mais eficiente. "
                "O resultado ótimo está disponível abaixo."
            )
            cmp_result["suboptimal_note"] = ""
            cmp_result["optimal_result"]  = None
            optimal_result = cmp_result

    best_result["suboptimal_note"] = suboptimal_note
    best_result["optimal_result"]  = optimal_result

    return best_result


# ── Write results back to xlsx ─────────────────────────────────────────────────

def write_results(filepath: str, route: dict, inv: dict):
    wb      = openpyxl.load_workbook(filepath)
    ws_calc = wb["⚜️ Relic Calculator"]

    thin  = Side(style="thin",   color="CCCCCC")
    med   = Side(style="medium", color="888888")
    def brd():   return Border(left=thin, right=thin, top=thin,  bottom=thin)
    def brd_m(): return Border(left=med,  right=med,  top=med,   bottom=med)
    def fl(h):   return PatternFill("solid", start_color=h, fgColor=h)
    def ca():    return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def la():    return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    target_level = inv["config"]["target_level"]

    def tier_dark(idx):
        if idx == 0:    return "0D1B2A"
        if idx <= 25:   return "B7950B"
        if idx <= 50:   return "922B21"
        if idx <= 75:   return "1A5276"
        return                 "212F3C"

    def tier_light(idx):
        if idx == 0:    return "EFEFEF"
        if idx <= 25:   return "FEF9E7"
        if idx <= 50:   return "FDEDEC"
        if idx <= 75:   return "EBF5FB"
        return                 "EAECEE"

    # ── State table ───────────────────────────────────────────────────────────
    for row in ws_calc.iter_rows(min_row=14, values_only=False):
        raw = str(row[1].value or "").strip().lstrip("⭐").strip()
        if raw not in ALL_RELICS:
            continue
        r         = row[0].row
        final_idx = route["final_levels"].get(raw, 0)
        final_sp  = route["final_specific"].get(raw, 0)
        orig_sp   = inv["relics"].get(raw, {}).get("specific_shards", 0)
        sp_used   = orig_sp - final_sp
        dark      = tier_dark(final_idx)
        light     = tier_light(final_idx)
        star, leg = idx_to_star_leg(final_idx)
        is_target = raw in route["targets"]
        is_inter  = raw in [inv["config"]["inter1"], inv["config"]["inter2"]]
        reached   = final_idx >= target_level

        for ci, val, extra_font, extra_fill in [
            (6, star,
             Font(name="Arial", color=dark, bold=True, size=9),
             fl(light)),
            (7, leg if final_idx > 0 else "—",
             Font(name="Arial", color=dark, size=9),
             fl(light)),
            (8, sp_used if sp_used > 0 else "—",
             Font(name="Arial", color="1A6B2A" if sp_used > 0 else "AAAAAA", size=9),
             fl(light)),
        ]:
            c = ws_calc.cell(r, ci)
            c.value = val; c.font = extra_font
            c.fill  = extra_fill; c.alignment = ca(); c.border = brd()

        c9 = ws_calc.cell(r, 9)
        if is_target and not reached:
            missing = shards_needed(final_idx, target_level)
            c9.value = f"-{missing}"
            c9.font  = Font(name="Arial", color="AA0000", bold=True, size=9)
        else:
            c9.value = "—"
            c9.font  = Font(name="Arial", color="AAAAAA", size=9)
        c9.fill = fl(light); c9.alignment = ca(); c9.border = brd()

        note = ""
        if is_target:
            note = "Target reached" if reached else "Target not reached"
        elif is_inter:
            note = "Used as chain relay"
        c10 = ws_calc.cell(r, 10)
        c10.value = note
        c10.font  = Font(name="Arial", size=8,
                         color="1A6B2A" if reached else "884400")
        c10.fill  = fl(light); c10.alignment = la(); c10.border = brd()

    # ── Route steps ───────────────────────────────────────────────────────────
    route_row = None
    for row in ws_calc.iter_rows(min_row=30, values_only=False):
        if str(row[0].value or "") == "Step":
            route_row = row[0].row + 1
            break

    if route_row:
        for clear_r in range(route_row, route_row + 200):
            row_empty = True
            for ci in range(1, 11):
                c = ws_calc.cell(clear_r, ci)
                if c.value is not None:
                    c.value = None
                    row_empty = False
            if clear_r > route_row + 10 and row_empty:
                break

        step_num = 1
        r        = route_row

        for step in route["steps"]:
            ws_calc.row_dimensions[r].height = 18

            if step["type"] == "swap":
                a_s, a_l = idx_to_star_leg(step["a_from"])
                b_s, b_l = idx_to_star_leg(step["b_to"])
                vals = [
                    "Hammer",
                    "Miracle Hammer SWAP",
                    f"{step['relic_a']}  <->  {step['relic_b']}",
                    f"{step['relic_a']}: {a_s}", a_l,
                    f"{step['relic_b']}: {b_s}", b_l,
                    "—", "—",
                    (f"{step['relic_b']} rises to {b_s} {b_l}"
                     f" | {step['relic_a']} drops to "
                     + " ".join(idx_to_star_leg(step["a_to"]))),
                ]
                bg_use = "FFF3E0"
                bold   = True
                color  = "8B4513"

            else:
                from_s, from_l = idx_to_star_leg(step["from"])
                to_s,   to_l   = idx_to_star_leg(step["to"])
                vals = [
                    step_num,
                    f"Develop  ->  {to_s} {to_l}",
                    step["relic"],
                    from_s, from_l,
                    to_s,   to_l,
                    step["sp_used"] if step["sp_used"] else "—",
                    step["u_used"]  if step["u_used"]  else "—",
                    "",
                ]
                step_num += 1
                bg_use = "F5F8FF" if step_num % 2 == 0 else "FFFFFF"
                bold   = False
                color  = "000000"

            for ci, val in enumerate(vals, 1):
                c = ws_calc.cell(r, ci)
                c.value     = val
                c.font      = Font(name="Arial", size=8, bold=bold, color=color)
                c.fill      = PatternFill("solid", start_color=bg_use, fgColor=bg_use)
                c.alignment = la() if ci in (2, 3, 10) else ca()
                c.border    = brd()
            r += 1

    # ── Summary ───────────────────────────────────────────────────────────────
    for row in ws_calc.iter_rows(min_row=50, values_only=False):
        if str(row[0].value or "").startswith("📊"):
            sr = row[0].row + 1
            for col, val in [
                (2, route["hammers_used"]),
                (4, route["universal_used"]),
                (6, route["hammers_used"]),
                (8, inv["config"]["hammers_avail"] - route["hammers_used"]),
            ]:
                c = ws_calc.cell(sr, col)
                c.value     = val
                c.font      = Font(name="Arial", color="1A6B2A", bold=True, size=11)
                c.fill      = PatternFill("solid", start_color="E6F4EA", fgColor="E6F4EA")
                c.alignment = ca()
                c.border    = brd_m()
            break

    wb.save(filepath)

    print(f"Route written to: {filepath}")
    print(f"  Targets:          {', '.join(route['targets'])}")
    develop_steps = [s for s in route["steps"] if s["type"] == "develop"]
    swap_steps    = [s for s in route["steps"] if s["type"] == "swap"]
    print(f"  Swaps (hammers):  {len(swap_steps)}")
    print(f"  Develop steps:    {len(develop_steps)}")
    print(f"  Hammers used:     {route['hammers_used']} / {inv['config']['hammers_avail']}")
    print(f"  Universal shards: {route['universal_used']} / {inv['universal_shards']}")
    if route.get("assignment"):
        for tname, mlist in sorted(route["assignment"].items()):
            for mname in mlist:
                print(f"  {mname}  ->  {tname}")
    if route.get("suboptimal_note"):
        print()
        print(route["suboptimal_note"])


# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    filepath = sys.argv[1] if len(sys.argv) > 1 else "Relic_Calculator_v2.xlsx"

    print(f"Reading:  {filepath}")
    inv = read_inventory(filepath)
    cfg = inv["config"]
    tgt_s, tgt_l = idx_to_star_leg(cfg["target_level"])

    print(f"Target:   {cfg['target_set']}  ->  {tgt_s} {tgt_l}")
    print(f"Mandatory relics: {cfg['inter1']}"
          + (f"  +  {cfg['inter2']}" if cfg["inter2"] else ""))
    print(f"Hammers:  {cfg['hammers_avail']}")
    print(f"Universal shards: {inv['universal_shards']}")
    print()

    route = compute_route(inv)

    print("Per-target results:")
    for t in route["targets"]:
        orig  = inv["relics"].get(t, {}).get("star_idx", 0)
        final = route["final_levels"].get(t, orig)
        os_, ol = idx_to_star_leg(orig)
        fs,  fl = idx_to_star_leg(final)
        goal    = "GOAL REACHED" if final >= cfg["target_level"] else ""
        print(f"  {t}: {os_} {ol}  ->  {fs} {fl}  {goal}")
    print()

    write_results(filepath, route, inv)

    import os
    os.startfile(os.path.abspath(filepath))
